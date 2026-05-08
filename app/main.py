import io
import csv
from datetime import datetime
from fastapi import FastAPI, Request, Depends, HTTPException, Form
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

from app.database import get_db, engine
from app import models, schemas, crud, auth
from app.models import Base
import app.schemas as schemas

Base.metadata.create_all(bind=engine)

app = FastAPI(title="Calculation World")
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


def get_current_user(request: Request, db: Session = Depends(get_db)):
    token = request.cookies.get("access_token")
    if not token:
        return None
    username = auth.decode_token(token)
    if not username:
        return None
    return crud.get_user_by_username(db, username)


@app.get("/", response_class=HTMLResponse)
def index(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if user:
        return RedirectResponse("/dashboard", status_code=302)
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    registered = request.query_params.get("registered")
    return templates.TemplateResponse("login.html", {"request": request, "registered": registered})


@app.post("/login")
def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = crud.authenticate_user(db, username, password)
    if not user:
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Invalid username or password"},
            status_code=400,
        )
    token = auth.create_access_token({"sub": user.username})
    resp = RedirectResponse("/dashboard", status_code=302)
    resp.set_cookie("access_token", token, httponly=True, max_age=3600)
    return resp


@app.get("/register", response_class=HTMLResponse)
def register_page(request: Request):
    return templates.TemplateResponse("register.html", {"request": request})


@app.post("/register")
def register(
    request: Request,
    username: str = Form(...),
    email: str    = Form(...),
    password: str = Form(...),
    db: Session   = Depends(get_db),
):
    if crud.get_user_by_username(db, username):
        return templates.TemplateResponse(
            "register.html",
            {"request": request, "error": "Username already taken"},
            status_code=400,
        )
    if crud.get_user_by_email(db, email):
        return templates.TemplateResponse(
            "register.html",
            {"request": request, "error": "Email already registered"},
            status_code=400,
        )
    crud.create_user(db, schemas.UserCreate(username=username, email=email, password=password))
    return RedirectResponse("/login?registered=1", status_code=302)


@app.get("/forgot-username", response_class=HTMLResponse)
def forgot_username_page(request: Request):
    return templates.TemplateResponse("forgot_username.html", {"request": request})


@app.post("/forgot-username", response_class=HTMLResponse)
def forgot_username(
    request: Request,
    email: str = Form(...),
    db: Session = Depends(get_db),
):
    user = crud.get_user_by_email(db, email)
    message = f"Your username is: {user.username}" if user else "No account found with that email."
    return templates.TemplateResponse(
        "forgot_username.html", {"request": request, "message": message}
    )


@app.get("/reset-password", response_class=HTMLResponse)
def reset_password_page(request: Request):
    return templates.TemplateResponse("reset_password.html", {"request": request})


@app.post("/reset-password", response_class=HTMLResponse)
def reset_password(
    request: Request,
    email: str        = Form(...),
    new_password: str = Form(...),
    db: Session       = Depends(get_db),
):
    user = crud.get_user_by_email(db, email)
    if not user:
        return templates.TemplateResponse(
            "reset_password.html",
            {"request": request, "error": "No account found with that email."},
        )
    crud.update_password(db, user, new_password)
    return templates.TemplateResponse(
        "reset_password.html",
        {"request": request, "message": "Password updated! You can now log in."},
    )


@app.get("/logout")
def logout():
    resp = RedirectResponse("/login", status_code=302)
    resp.delete_cookie("access_token")
    return resp


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    recent = crud.get_history(db, user.id, limit=5)
    stats  = crud.get_stats(db, user.id)
    return templates.TemplateResponse(
        "dashboard.html",
        {"request": request, "user": user, "recent": recent, "stats": stats},
    )


@app.get("/calculator", response_class=HTMLResponse)
def calculator_page(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("calculator.html", {"request": request, "user": user})


@app.post("/calculate", response_class=HTMLResponse)
def calculate(
    request: Request,
    a: float       = Form(...),
    b: float       = Form(...),
    operation: str = Form(...),
    db: Session    = Depends(get_db),
):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)

    OPS = {
        "add":      (lambda x, y: x + y,    "+"),
        "subtract": (lambda x, y: x - y,    "-"),
        "multiply": (lambda x, y: x * y,    "x"),
        "divide":   (lambda x, y: x / y,    "/"),
        "power":    (lambda x, y: x ** y,   "^"),
        "modulus":  (lambda x, y: x % y,    "mod"),
        "sqrt_a":   (lambda x, y: x ** 0.5, "sqrt"),
    }

    result = None
    error  = None

    if operation not in OPS:
        error = "Unknown operation."
    elif operation in ("divide", "modulus") and b == 0:
        error = "Cannot divide by zero."
    else:
        try:
            fn, _ = OPS[operation]
            result = fn(a, b)
            crud.create_calculation(
                db,
                schemas.CalculationCreate(
                    user_id=user.id, operand_a=a,
                    operand_b=b, operation=operation, result=result,
                ),
            )
        except Exception as e:
            error = str(e)

    return templates.TemplateResponse(
        "calculator.html",
        {"request": request, "user": user,
         "a": a, "b": b, "operation": operation,
         "result": result, "error": error},
    )


@app.get("/history", response_class=HTMLResponse)
def history_page(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    records = crud.get_history(db, user.id)
    stats   = crud.get_stats(db, user.id)
    return templates.TemplateResponse(
        "history.html",
        {"request": request, "user": user, "records": records, "stats": stats},
    )


@app.get("/history/export/csv")
def export_csv(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    records = crud.get_history(db, user.id)
    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(["ID", "Operation", "A", "B", "Result", "Timestamp"])
    for r in records:
        w.writerow([r.id, r.operation, r.operand_a, r.operand_b, r.result,
                    r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else ""])
    buf.seek(0)
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=calc_history.csv"},
    )


@app.get("/history/export/pdf")
def export_pdf(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    records = crud.get_history(db, user.id)
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=letter)
    styl = getSampleStyleSheet()
    elems = [
        Paragraph(f"Calculation History - {user.username}", styl["Title"]),
        Paragraph(f"Exported: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styl["Normal"]),
        Spacer(1, 20),
    ]
    data = [["#", "Operation", "A", "B", "Result", "Date"]]
    for r in records:
        data.append([str(r.id), r.operation, str(r.operand_a), str(r.operand_b),
                     f"{r.result:.4f}",
                     r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else ""])
    t = Table(data, colWidths=[30, 70, 70, 70, 80, 120])
    t.setStyle(TableStyle([
        ("BACKGROUND",     (0,0), (-1,0), colors.HexColor("#6c63ff")),
        ("TEXTCOLOR",      (0,0), (-1,0), colors.white),
        ("FONTNAME",       (0,0), (-1,0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#f0eeff"), colors.white]),
        ("GRID",           (0,0), (-1,-1), 0.5, colors.HexColor("#cccccc")),
        ("FONTSIZE",       (0,0), (-1,-1), 9),
    ]))
    elems.append(t)
    doc.build(elems)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=calc_history.pdf"})


@app.get("/history/export/excel")
def export_excel(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    records = crud.get_history(db, user.id)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "History"
    headers = ["ID", "Operation", "A", "B", "Result", "Timestamp"]
    hf = Font(bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor="6C63FF")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hb
        cell.alignment = Alignment(horizontal="center")
    for i, r in enumerate(records, 2):
        ws.cell(row=i, column=1, value=r.id)
        ws.cell(row=i, column=2, value=r.operation)
        ws.cell(row=i, column=3, value=r.operand_a)
        ws.cell(row=i, column=4, value=r.operand_b)
        ws.cell(row=i, column=5, value=round(r.result, 4))
        ws.cell(row=i, column=6, value=r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=calc_history.xlsx"})


@app.get("/profile", response_class=HTMLResponse)
def profile_page(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("profile.html", {"request": request, "user": user})


@app.post("/profile", response_class=HTMLResponse)
def update_profile(
    request: Request,
    email: str            = Form(...),
    current_password: str = Form(...),
    new_password: str     = Form(""),
    db: Session           = Depends(get_db),
):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not auth.verify_password(current_password, user.hashed_password):
        return templates.TemplateResponse(
            "profile.html",
            {"request": request, "user": user, "error": "Current password is incorrect."},
        )
    crud.update_email(db, user, email)
    if new_password.strip():
        crud.update_password(db, user, new_password)
    return templates.TemplateResponse(
        "profile.html",
        {"request": request, "user": user, "message": "Profile updated successfully!"},
    )


@app.get("/api/history", response_model=list[schemas.CalculationOut])
def api_history(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return crud.get_history(db, user.id)


@app.post("/api/calculate", response_model=schemas.CalculationOut)
def api_calculate(
    payload: schemas.CalculationIn,
    request: Request,
    db: Session = Depends(get_db),
):
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    OPS = {
        "add":      lambda a, b: a + b,
        "subtract": lambda a, b: a - b,
        "multiply": lambda a, b: a * b,
        "divide":   lambda a, b: a / b,
        "power":    lambda a, b: a ** b,
        "modulus":  lambda a, b: a % b,
        "sqrt_a":   lambda a, b: a ** 0.5,
    }
    if payload.operation not in OPS:
        raise HTTPException(status_code=400, detail="Unknown operation")
    if payload.operation in ("divide", "modulus") and payload.operand_b == 0:
        raise HTTPException(status_code=400, detail="Cannot divide by zero")
    result = OPS[payload.operation](payload.operand_a, payload.operand_b)
    return crud.create_calculation(db, schemas.CalculationCreate(
        user_id=user.id, operand_a=payload.operand_a,
        operand_b=payload.operand_b, operation=payload.operation, result=result,
    ))
